
from __future__ import annotations

import math
import os
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, f1_score
from sklearn.pipeline import Pipeline
from text_preprocessing import normalize_text


print("RUNNING FILE:", Path(__file__).resolve())

TRAIN_PATH = "data/processed/train.csv"
TEST_PATH = "data/processed/test.csv"

# Ausgabe: genau 1 Datei
OUT_XLSX = "data/processed/learning_curve_report.xlsx"

# Optional: falls alte CSVs existieren -> löschen
DELETE_OLD_CSVS = True
OLD_FILES = [
    "data/processed/learning_curve.csv",
    "data/processed/learning_curve_summary.csv",
]

# 10 Punkte (du kannst das ändern)
POINTS = [500, 1000, 2000, 4000, 8000, 12000, 16000, 20000, 26000, 33600]


def train_and_eval(train_df: pd.DataFrame, test_df: pd.DataFrame) -> dict:
    model = Pipeline(
        [
            ("tfidf", TfidfVectorizer(
                preprocessor=normalize_text,
                ngram_range=(1, 2),
                min_df=2,
                max_df=0.95
            )),
            ("clf", LogisticRegression(max_iter=2000)),
        ]
    )

    model.fit(train_df["text"], train_df["label"])

    # Gesamt
    y_pred = model.predict(test_df["text"])
    acc_all = accuracy_score(test_df["label"], y_pred)
    f1_all = f1_score(test_df["label"], y_pred)

    out = {"acc_all": acc_all, "f1_all": f1_all}

    # Pro Sprache
    for lang in ["de", "en"]:
        sub = test_df[test_df["lang"] == lang]
        y_sub_pred = model.predict(sub["text"])
        out[f"acc_{lang}"] = accuracy_score(sub["label"], y_sub_pred)
        out[f"f1_{lang}"] = f1_score(sub["label"], y_sub_pred)

    return out


def stratified_sample(df: pd.DataFrame, n: int, seed: int = 42) -> pd.DataFrame:
    """Zieht ca. n Zeilen, aber hält lang+label-Verteilung stabil."""
    strat = df["lang"].astype(str) + "_" + df["label"].astype(str)
    total = len(df)

    parts = []
    for _, g in df.groupby(strat):
        share = len(g) / total
        k = max(1, math.floor(n * share))
        k = min(k, len(g))
        parts.append(g.sample(n=k, random_state=seed))

    sampled = pd.concat(parts, ignore_index=True)

    # auf exakt n trimmen (falls leicht drüber)
    if len(sampled) > n:
        sampled = sampled.sample(n=n, random_state=seed).reset_index(drop=True)
    else:
        sampled = sampled.sample(frac=1.0, random_state=seed).reset_index(drop=True)

    return sampled


def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def build_excel(df_curve: pd.DataFrame, test_df: pd.DataFrame, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lernkurve"

    title = "Lernkurve (F1) – Gesamt vs. Deutsch vs. Englisch"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    # Header
    headers = ["train_size", "f1_all", "f1_de", "f1_en", "acc_all", "acc_de", "acc_en"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for col_idx, _h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    # Daten
    for _, r in df_curve.iterrows():
        ws.append(
            [
                int(r["train_size"]),
                float(r["f1_all"]),
                float(r["f1_de"]),
                float(r["f1_en"]),
                float(r["acc_all"]),
                float(r["acc_de"]),
                float(r["acc_en"]),
            ]
        )

    # Format
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = "0.0000"

    ws.freeze_panes = "A3"
    autosize_columns(ws)

    # Summary rechts daneben
    start_col = 9  # Spalte I
    ws.cell(row=2, column=start_col, value="Summary").font = Font(bold=True, size=12)

    final = df_curve.iloc[-1]
    summary_rows = [
        ("Final train_size", int(final["train_size"])),
        ("Final F1 (Gesamt)", float(final["f1_all"])),
        ("Final F1 (DE)", float(final["f1_de"])),
        ("Final F1 (EN)", float(final["f1_en"])),
        ("Final Acc (Gesamt)", float(final["acc_all"])),
        ("Final Acc (DE)", float(final["acc_de"])),
        ("Final Acc (EN)", float(final["acc_en"])),
        ("Test n (Gesamt)", int(len(test_df))),
        ("Test n (DE)", int((test_df["lang"] == "de").sum())),
        ("Test n (EN)", int((test_df["lang"] == "en").sum())),
    ]

    r0 = 3
    ws.cell(row=r0, column=start_col, value="Metric").font = Font(bold=True)
    ws.cell(row=r0, column=start_col + 1, value="Value").font = Font(bold=True)

    for i, (k, v) in enumerate(summary_rows, start=1):
        ws.cell(row=r0 + i, column=start_col, value=k)
        cell_v = ws.cell(row=r0 + i, column=start_col + 1, value=v)
        if isinstance(v, float):
            cell_v.number_format = "0.0000"

    # ===== Chart =====
    chart = LineChart()
    chart.title = "Lernkurve (F1) – Gesamt vs. Deutsch vs. Englisch"
    chart.style = 2
    chart.height = 10
    chart.width = 24

    chart.y_axis.title = "F1-Score"
    chart.x_axis.title = "Trainingsgröße (n)"

    # Y-Achse Skala + Ticks
    chart.y_axis.scaling.min = 0.50
    chart.y_axis.scaling.max = 0.95
    chart.y_axis.majorUnit = 0.05
    chart.y_axis.number_format = "0.00"

    # X-Achse: sinnvolle Ablesbarkeit
    chart.x_axis.majorUnit = 5000
    chart.x_axis.number_format = "0"

    # Achsen-Zahlen sichtbar machen
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.tickLblPos = "nextTo"

    # X-Gridlines aus (optional)
    chart.x_axis.majorGridlines = None

    # Legende oben
    chart.legend.position = "t"
    chart.legend.overlay = False

        # Achsen + Zahlen erzwingen (sonst manchmal ohne Tick-Labels)
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"

    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"

    chart.x_axis.number_format = "0"
    chart.y_axis.number_format = "0.00"

    # Datenbereich (F1-Spalten B-D)
    min_row = 2
    max_row = 2 + len(df_curve)
    data = Reference(ws, min_col=2, max_col=4, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Linien dicker + Marker an (besser lesbar)
    for s in chart.series:
        s.marker.symbol = "circle"
        s.marker.size = 6
        # openpyxl: Breite über properties (funktioniert stabil)
        try:
            s.graphicalProperties.line.width = 22000  # etwas dicker
        except Exception:
            pass

    ws.add_chart(chart, "I2")

    # Interpretation
    ws["I14"] = "Interpretation (kurz):"
    ws["I14"].font = Font(bold=True)
    ws["I15"] = "EN ist stabiler/höher als DE, weil viel mehr EN-Trainingsdaten vorhanden sind."
    ws["I16"] = "DE ist schwieriger (weniger Daten + andere Domäne/Stil), daher geringere F1-Werte."

    autosize_columns(ws)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    # Optional alte CSVs löschen
    if DELETE_OLD_CSVS:
        for f in OLD_FILES:
            try:
                os.remove(f)
            except FileNotFoundError:
                pass

    train_df = pd.read_csv(TRAIN_PATH)
    test_df = pd.read_csv(TEST_PATH)

    max_n = len(train_df)
    rows = []

    for n in POINTS:
        n = min(n, max_n)
        sub = stratified_sample(train_df, n=n, seed=42)
        m = train_and_eval(sub, test_df)

        row = {
            "train_size": len(sub),
            "f1_all": m["f1_all"],
            "f1_de": m["f1_de"],
            "f1_en": m["f1_en"],
            "acc_all": m["acc_all"],
            "acc_de": m["acc_de"],
            "acc_en": m["acc_en"],
        }
        rows.append(row)
        print(
            f"✅ n={row['train_size']} | f1_all={row['f1_all']:.4f} | "
            f"f1_de={row['f1_de']:.4f} | f1_en={row['f1_en']:.4f}"
        )

    df_curve = pd.DataFrame(rows)
    build_excel(df_curve, test_df, OUT_XLSX)
    print(f"\n✅ Fertig: {OUT_XLSX}")

    # Optional automatisch öffnen (Windows)
    try:
        os.startfile(OUT_XLSX)  # noqa: S606
    except Exception:
        pass


if __name__ == "__main__":
    main()
